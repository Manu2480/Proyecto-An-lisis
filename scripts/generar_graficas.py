import os
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import seaborn as sns
from openpyxl import load_workbook

warnings.filterwarnings("ignore")
sns.set_theme(style="whitegrid", font_scale=0.95)

RUTA_EXCEL = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "docs", "DatosPruebas2026_1_filled.xlsx"
)
DIR_SALIDA = os.path.join(os.path.dirname(os.path.dirname(__file__)), "outputs", "plots")
os.makedirs(DIR_SALIDA, exist_ok=True)

HOJAS_INTERES = ["10A-Elementos", "15B-Elementos", "20A-Elementos"]

# column offsets per k (0-indexed) for each strategy block
# For each k=2..5: QNodes(Partición, Pérdida, Tiempo) + Geometric(Partición, Pérdida, Tiempo) = 6 cols per k
BASE = 3  # data columns start at index 3 (after #Prueba, Alcance, Mecanismo)
K_VALS = [2, 3, 4, 5]
ESTRATEGIAS = ["QNodes", "Geometric"]

def parse_sheet(ws):
    rows = []
    for r in range(6, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if vals[0] is None:
            continue
        prueba = vals[0]
        subsistema = vals[1] if vals[1] is not None else ""
        mecanismo = vals[2] if vals[2] is not None else ""
        tamano = len(str(subsistema)) if subsistema else 0
        for ki, k in enumerate(K_VALS):
            offset = BASE + ki * 6
            for si, strat in enumerate(ESTRATEGIAS):
                so = offset + si * 3
                particion = vals[so] if so < len(vals) else None
                perdida = vals[so + 1] if so + 1 < len(vals) else None
                tiempo = vals[so + 2] if so + 2 < len(vals) else None
                if perdida is not None and tiempo is not None:
                    rows.append({
                        "prueba": prueba,
                        "subsistema": subsistema,
                        "mecanismo": mecanismo,
                        "tamano": tamano,
                        "k": k,
                        "estrategia": strat,
                        "particion": str(particion) if particion else "",
                        "perdida": float(perdida),
                        "tiempo": float(tiempo),
                    })
    return pd.DataFrame(rows)


def grafico1_tiempo_vs_tamano(df, nombre_hoja):
    """Gráfico 1: Tiempo de ejecución vs Tamaño para K-Particiones (Geometric), facetado por k."""
    df_kp = df[df["estrategia"] == "Geometric"].copy()
    if df_kp.empty:
        print(f"  [AVISO] {nombre_hoja}: Sin datos Geometric para Gráfico 1")
        return
    df_kp["tamano"] = df_kp["tamano"].astype(int)
    df_kp["k"] = df_kp["k"].astype(int)
    # Aggregate by (tamano, k) for cleaner lines
    agg = df_kp.groupby(["tamano", "k"], as_index=False)["tiempo"].mean()
    fig, axes = plt.subplots(2, 2, figsize=(10, 8), sharey=False)
    axes = axes.flatten()
    for idx, k in enumerate([2, 3, 4, 5]):
        ax = axes[idx]
        sub = agg[agg["k"] == k].sort_values("tamano")
        ax.plot(sub["tamano"], sub["tiempo"], marker="o", color="steelblue", linewidth=1.5)
        ax.set_title(f"k = {k}", fontweight="bold")
        ax.set_xlabel("Tamaño del subsistema")
        ax.set_ylabel("Tiempo de ejecución (ms)")
        ax.yaxis.set_major_formatter(mticker.ScalarFormatter(useMathText=True))
        ax.ticklabel_format(style="sci", axis="y", scilimits=(0,0))
        ax.grid(True, alpha=0.3)
    fig.suptitle(f"{nombre_hoja} – Rendimiento Temporal (K-Particiones / Geometric)", fontweight="bold", y=1.02)
    plt.tight_layout()
    ruta = os.path.join(DIR_SALIDA, f"{nombre_hoja}_grafico_1.png")
    fig.savefig(ruta, dpi=200, bbox_inches="tight")
    plt.close(fig)
    print(f"  -> Guardado: {ruta}")


def grafico2_boxplot_perdida(df, nombre_hoja):
    """Gráfico 2: Boxplot de pérdida (EMD) agrupado por subsistema y estrategia."""
    df = df.copy()
    df["subsistema_label"] = df["subsistema"].astype(str) + " (n=" + df["tamano"].astype(str) + ")"
    orden = sorted(df["subsistema_label"].unique(), key=lambda x: int(x.split("n=")[1].rstrip(")")))
    plt.figure(figsize=(max(10, len(orden) * 0.6), 5))
    ax = sns.boxplot(
        data=df,
        x="subsistema_label", y="perdida",
        hue="estrategia",
        order=orden,
        palette={"QNodes": "#3498db", "Geometric": "#e67e22"},
        linewidth=0.8,
        fliersize=2,
    )
    ax.set_xlabel("Subsistema (tamaño)")
    ax.set_ylabel("Pérdida (EMD)")
    ax.set_title(f"{nombre_hoja} – Comparativa de Pérdida por Estrategia", fontweight="bold")
    ax.legend(title="Estrategia")
    ax.grid(True, alpha=0.3)
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    ruta = os.path.join(DIR_SALIDA, f"{nombre_hoja}_grafico_2.png")
    plt.savefig(ruta, dpi=200, bbox_inches="tight")
    plt.close()
    print(f"  -> Guardado: {ruta}")

    # Análisis: ¿se halla la misma k-partición óptima entre estrategias para un mismo subsistema?
    print(f"\n  --- Análisis Gráfico 2 ({nombre_hoja}) ---")
    for (sub, tam), grp in df.groupby(["subsistema", "tamano"]):
        for k in K_VALS:
            gk = grp[grp["k"] == k]
            if gk.empty or len(gk) < 2:
                continue
            qnodes_part = gk[gk["estrategia"] == "QNodes"]["particion"].values
            geo_part = gk[gk["estrategia"] == "Geometric"]["particion"].values
            if len(qnodes_part) == 0 or len(geo_part) == 0:
                continue
            # Normalize partition strings for comparison
            qp = set(qnodes_part[0].strip())
            gp = set(geo_part[0].strip())
            # Quick check if partitions are structurally similar
            same = qp == gp
            status = "SÍ" if same else "NO"
            print(f"    k={k}, subsistema={sub} (n={tam}): ¿misma partición óptima? {status}")


def grafico3_desviacion_qnodes(df, nombre_hoja):
    """Gráfico 3: Desviación de pérdida respecto a QNodes (baseline)."""
    qnodes = df[df["estrategia"] == "QNodes"].copy()
    geo = df[df["estrategia"] == "Geometric"].copy()
    merged = pd.merge(
        qnodes, geo,
        on=["subsistema", "mecanismo", "prueba", "k", "tamano"],
        suffixes=("_qnodes", "_geo"),
        how="inner"
    )
    if merged.empty:
        print(f"  [AVISO] {nombre_hoja}: Sin datos para Gráfico 3")
        return
    merged["delta_perdida"] = merged["perdida_geo"] - merged["perdida_qnodes"]
    merged["subsistema_label"] = merged["subsistema"].astype(str) + " (n=" + merged["tamano"].astype(str) + ")"
    orden = sorted(merged["subsistema_label"].unique(), key=lambda x: int(x.split("n=")[1].rstrip(")")))

    plt.figure(figsize=(max(10, len(orden) * 0.5), 5))
    colors = merged["delta_perdida"].apply(lambda x: "#2ecc71" if x <= 0 else "#e74c3c")
    ax = sns.barplot(
        data=merged,
        x="subsistema_label", y="delta_perdida",
        hue="k", palette="viridis",
        order=orden,
        ci=None,
    )
    ax.axhline(0, color="black", linewidth=0.8)
    ax.set_xlabel("Subsistema (tamaño)")
    ax.set_ylabel("Δ Pérdida (Geometric − QNodes)")
    ax.set_title(f"{nombre_hoja} – Variación de Pérdida respecto a QNodes (baseline)", fontweight="bold")
    ax.legend(title="k", bbox_to_anchor=(1.02, 1), loc="upper left")
    ax.grid(True, alpha=0.3)
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    ruta = os.path.join(DIR_SALIDA, f"{nombre_hoja}_grafico_3.png")
    plt.savefig(ruta, dpi=200, bbox_inches="tight")
    plt.close()
    print(f"  -> Guardado: {ruta}")


def main():
    wb = load_workbook(RUTA_EXCEL, data_only=True)
    for nombre_hoja in HOJAS_INTERES:
        if nombre_hoja not in wb.sheetnames:
            print(f"[SKIP] Hoja '{nombre_hoja}' no encontrada.")
            continue
        print(f"\n{'='*60}")
        print(f"Procesando: {nombre_hoja}")
        print(f"{'='*60}")
        ws = wb[nombre_hoja]
        df = parse_sheet(ws)
        print(f"  Filas parseadas: {len(df)}")
        if df.empty:
            print(f"  [SKIP] Sin datos.")
            continue

        grafico1_tiempo_vs_tamano(df, nombre_hoja)
        grafico2_boxplot_perdida(df, nombre_hoja)
        grafico3_desviacion_qnodes(df, nombre_hoja)

    print(f"\n¡Todos los gráficos generados en: {DIR_SALIDA}/")


if __name__ == "__main__":
    main()
