import pandas as pd
import openpyxl

csv_path = r'C:\Users\User\Desktop\Proyecto-An-lisis\GeoMIP\data\results\comparativa\comparativa_long.csv'
xlsx_path = r'C:\Users\User\Desktop\Proyecto-An-lisis\GeoMIP\data\results\DatosPruebas2026_1 (1).xlsx'

df = pd.read_csv(csv_path)

n_to_sheet = {
    10: '10A-Elementos',
    15: '15B-Elementos',
    20: '20A-Elementos',
    22: '22A-Elementos',
    25: '25A-Elementos',
}

k_col_mapping = {
    2: {'QNodos_perdida': 5, 'QNodos_tiempo': 6, 'Geometric_perdida': 8, 'Geometric_tiempo': 9},
    3: {'QNodos_perdida': 11, 'QNodos_tiempo': 12, 'Geometric_perdida': 14, 'Geometric_tiempo': 15},
    4: {'QNodos_perdida': 17, 'QNodos_tiempo': 18, 'Geometric_perdida': 20, 'Geometric_tiempo': 21},
    5: {'QNodos_perdida': 23, 'QNodos_tiempo': 24, 'Geometric_perdida': 26, 'Geometric_tiempo': 27},
}

mode_to_strategy = {
    'Exacto': 'QNodos',
    'Rapido_MCTS': 'Geometric',
}

wb = openpyxl.load_workbook(xlsx_path)

sheet_map = {name.strip(): name for name in wb.sheetnames}

for n_val in [25, 22, 20, 15, 10]:
    target = n_to_sheet.get(n_val)
    actual_name = sheet_map.get(target)
    if actual_name is None:
        print(f'Sheet matching "{target}" not found for n={n_val}')
        continue

    ws = wb[actual_name]
    print(f'\n=== Processing {actual_name} (n={n_val}) ===')

    n_data = df[df['n'] == n_val]
    print(f'  CSV rows for n={n_val}: {len(n_data)}')
    print(f'  Max row in sheet: {ws.max_row}')

    max_row = ws.max_row
    filled_count = 0
    row_match_count = 0

    for row_idx in range(6, max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1).value
        cell_b = ws.cell(row=row_idx, column=2).value
        cell_c = ws.cell(row=row_idx, column=3).value

        if cell_b is None and cell_c is None:
            continue

        b_val = str(cell_b).strip() if cell_b else ''
        c_val = str(cell_c).strip() if cell_c else ''

        if not b_val or not c_val:
            continue

        prueba_num = None
        if cell_a is not None:
            a_str = str(cell_a)
            if a_str.startswith('=ROW(A'):
                try:
                    prueba_num = int(a_str.replace('=ROW(A', '').replace(')', ''))
                except:
                    pass

        if prueba_num is None:
            continue

        matching_rows = n_data[
            (n_data['#Prueba'] == prueba_num) &
            (n_data['Purview'].str.strip() == b_val) &
            (n_data['Mecanismo'].str.strip() == c_val)
        ]

        if len(matching_rows) == 0:
            continue

        row_match_count += 1

        for k_val in [2, 3, 4, 5]:
            for modo_name, strategy_name in mode_to_strategy.items():
                mode_rows = matching_rows[
                    (matching_rows['modo'] == modo_name) &
                    (matching_rows['k'] == k_val)
                ]

                if len(mode_rows) > 0:
                    row_data = mode_rows.iloc[0]

                    perdida_col = k_col_mapping[k_val][f'{strategy_name}_perdida']
                    tiempo_col = k_col_mapping[k_val][f'{strategy_name}_tiempo']

                    ws.cell(row=row_idx, column=perdida_col).value = float(row_data['perdida'])
                    ws.cell(row=row_idx, column=tiempo_col).value = float(row_data['tiempo_ms'])
                    filled_count += 1

    print(f'  Rows matched: {row_match_count}')
    print(f'  Cells filled: {filled_count}')
    if row_match_count > 0:
        total_possible = row_match_count * len([2,3,4,5]) * len(mode_to_strategy)
        print(f'  Expected max cells: {total_possible}')

output_path = r'C:\Users\User\Desktop\Proyecto-An-lisis\GeoMIP\data\results\DatosPruebas2026_1_filled.xlsx'
wb.save(output_path)
print(f'\nSaved to: {output_path}')
print('Done!')
